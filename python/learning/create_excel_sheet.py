import openpyxl

def create_excel_sheet(file_path):
    # Create a new workbook and select the active sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Define the dictionary with headers as keys and subheaders as values
    data_dict = {
        "Header 1": {"Subheader 1.1": 11, "Subheader 1.2": 12, "Subheader 1.3": 13},
        "Header 2": "hello",
        "Header 3": "hi",
        "Header 4": "value",
    }

    # Extract headers and subheaders from the dictionary
    headers = list(data_dict.keys())
    subheaders = {header: list(subheader_dict.keys()) for header, subheader_dict in data_dict.items()}

    # Write headers and merge cells for subheaders
    current_col = 1
    for header in headers:
        subheader_list = subheaders.get(header, [])
        header_col_span = max(len(subheader_list), 1)  # Ensure a minimum column span of 1
        sheet.merge_cells(start_row=1, start_column=current_col, end_row=1, end_column=current_col + header_col_span - 1)
        sheet.cell(row=1, column=current_col, value=header)
        current_col += header_col_span

    # Write subheaders
    current_col = 1
    for header, subheader_list in subheaders.items():
        for subheader in subheader_list:
            sheet.cell(row=2, column=current_col, value=subheader)
            current_col += 1

    # Insert sample values from the dictionary
    values = []
    for header in headers:
        subheader_dict = data_dict.get(header, {})
        row_values = [subheader_dict.get(subheader, None) for subheader in subheaders[header]]
        values.append(row_values)

    # Write values to the sheet
    for row_num, row_values in enumerate(values, 3):
        for col_num, value in enumerate(row_values, 1):
            sheet.cell(row=row_num, column=col_num, value=value)

    # Save the workbook to the specified file path
    workbook.save(file_path)
    print(f"Excel sheet created and saved at: {file_path}")

# Specify the file path where you want to save the Excel sheet
file_path = "example_excel_sheet_with_dict_values.xlsx"

# Call the function to create and save the Excel sheet
create_excel_sheet(file_path)
